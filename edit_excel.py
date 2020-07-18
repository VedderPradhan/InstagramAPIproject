from get_comments_list import getCommentIdList
from get_comment_info import getCommentatorAndComment

from openpyxl.styles import PatternFill
import openpyxl
import re

##commentIDs = getCommentIdList()

fileName = 'TonerGA.xlsx'
wb = openpyxl.load_workbook(filename = fileName)
ws1 = wb.active

lastRow = 794

for row in range(lastRow):
    cell = 'B' + str(row + 1)
    cellValue = str(ws1[cell].value)
    cellValue1 = cellValue.replace("\n","")
    print(cell)
    detectCon = re.search('(@\S*).*(@\S*).*(@\S*)(\s{1}|$)', cellValue1)
    #print (detectCon)
    if not detectCon:
        ws1[cell].fill = PatternFill("solid", fgColor="F71D1D")
    

wb.save(fileName)
    

    



